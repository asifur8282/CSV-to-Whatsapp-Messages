"""
File Handler Module
Handles Excel and CSV file operations
"""

import os
from pathlib import Path
import openpyxl
import pandas as pd
from src.user_input import UserInputHandler


class FileHandler:
    """Handles Excel and CSV file operations"""
    
    def __init__(self):
        self.user_input = UserInputHandler()
    
    def handle_file_input(self) -> list:
        """
        Handle file-based input for phone numbers
        
        Returns:
            list: List of phone numbers from the file
        """
        # Get file type
        file_type = self._get_file_type()
        
        # Get file path
        file_path = self._get_file_path(file_type)
        
        if file_type == "excel":
            return self._handle_excel_file(file_path)
        else:
            return self._handle_csv_file(file_path)
    
    @staticmethod
    def _get_file_type() -> str:
        """
        Ask user which file type they have
        
        Returns:
            str: 'excel' or 'csv'
        """
        print("\n📁 File Type Selection")
        
        while True:
            file_type = input("❓ Do you have an Excel or CSV file? (excel/csv): ").strip().lower()
            if file_type in ['excel', 'csv', 'e', 'c']:
                return 'excel' if file_type in ['excel', 'e'] else 'csv'
            print("   ⚠️  Please enter 'excel' or 'csv'")
    
    @staticmethod
    def _get_file_path(file_type: str) -> str:
        """
        Get file path from user and validate it
        
        Args:
            file_type: 'excel' or 'csv'
            
        Returns:
            str: Valid file path
        """
        extensions = ('.xlsx', '.xls', '.csv') if file_type == "excel" else ('.csv',)
        
        while True:
            file_path = input(f"\n❓ Enter the {file_type.upper()} file path: ").strip()
            
            # Handle quoted paths
            file_path = file_path.strip('"\'')
            
            # Expand user path
            file_path = os.path.expanduser(file_path)
            
            # Check if file exists
            if not os.path.exists(file_path):
                print(f"   ⚠️  File not found: {file_path}")
                continue
            
            # Check extension
            if not file_path.lower().endswith(extensions):
                print(f"   ⚠️  Invalid file extension. Expected {extensions}")
                continue
            
            return file_path
    
    def _handle_excel_file(self, file_path: str) -> list:
        """
        Handle Excel file and extract phone numbers
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            list: List of phone numbers
        """
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path)
            sheets = wb.sheetnames
            
            if not sheets:
                print("   ⚠️  Excel file has no sheets")
                return []
            
            # Let user select sheet if multiple sheets exist
            if len(sheets) > 1:
                print(f"\n📄 Available sheets: {', '.join(sheets)}")
                selected_sheet = self._select_sheet(sheets)
            else:
                selected_sheet = sheets[0]
            
            ws = wb[selected_sheet]
            
            # Find column with phone numbers
            column_letter, column_index = self._find_phone_column(ws)
            
            if not column_letter:
                print("   ⚠️  No column with 'Number' header found")
                return []
            
            # Extract phone numbers
            recipients = self._extract_phone_numbers_from_column(ws, column_index)
            
            if not recipients:
                print("   ⚠️  No phone numbers found in the selected column")
                return []
            
            print(f"\n✅ Found {len(recipients)} phone numbers in column {column_letter}")
            
            # Ask if user wants all or up to a specific row
            use_all = self.user_input.ask_yes_no("Use all numbers?")
            
            if not use_all:
                limit_row = self.user_input.get_integer_input(
                    "Enter the row number to limit to",
                    min_value=1,
                    max_value=len(recipients)
                )
                recipients = recipients[:limit_row]
                print(f"✅ Limited to {len(recipients)} numbers")
            
            return recipients
        
        except Exception as e:
            print(f"   ❌ Error processing Excel file: {str(e)}")
            return []
    
    def _handle_csv_file(self, file_path: str) -> list:
        """
        Handle CSV file and extract phone numbers
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            list: List of phone numbers
        """
        try:
            # Read CSV file
            df = pd.read_csv(file_path)
            
            # Find column with phone numbers
            column_name = self._find_csv_phone_column(df)
            
            if not column_name:
                print("   ⚠️  No column with 'Number' header found")
                return []
            
            # Extract phone numbers
            recipients = [str(num).strip() for num in df[column_name].astype(str) if str(num).strip()]
            
            if not recipients:
                print("   ⚠️  No phone numbers found in the selected column")
                return []
            
            print(f"\n✅ Found {len(recipients)} phone numbers in column '{column_name}'")
            
            # Ask if user wants all or up to a specific row
            use_all = self.user_input.ask_yes_no("Use all numbers?")
            
            if not use_all:
                limit_row = self.user_input.get_integer_input(
                    "Enter the row number to limit to",
                    min_value=1,
                    max_value=len(recipients)
                )
                recipients = recipients[:limit_row]
                print(f"✅ Limited to {len(recipients)} numbers")
            
            return recipients
        
        except Exception as e:
            print(f"   ❌ Error processing CSV file: {str(e)}")
            return []
    
    @staticmethod
    def _select_sheet(sheets: list) -> str:
        """
        Let user select a sheet
        
        Args:
            sheets: List of sheet names
            
        Returns:
            str: Selected sheet name
        """
        while True:
            selection = input("\n❓ Enter sheet name: ").strip()
            if selection in sheets:
                return selection
            print(f"   ⚠️  Sheet not found. Available: {', '.join(sheets)}")
    
    @staticmethod
    def _find_phone_column(ws) -> tuple:
        """
        Find column with phone numbers in Excel sheet
        
        Args:
            ws: Worksheet object
            
        Returns:
            tuple: (column_letter, column_index) or (None, None)
        """
        phone_keywords = ['number', 'phone', 'mobile', 'contact', 'tel', 'telephone', 'no.', 'no']
        
        for col in ws.iter_cols(min_row=1, max_row=1):
            cell = col[0]
            header = str(cell.value).lower() if cell.value else ""
            
            for keyword in phone_keywords:
                if keyword in header:
                    col_letter = cell.column_letter
                    col_index = cell.column
                    
                    # Confirm with user
                    confirm = input(f"\n❓ Use column {col_letter} ('{cell.value}')? (yes/no): ").strip().lower()
                    if confirm in ['yes', 'y']:
                        return col_letter, col_index
        
        return None, None
    
    @staticmethod
    def _find_csv_phone_column(df) -> str:
        """
        Find column with phone numbers in CSV file
        
        Args:
            df: Pandas DataFrame
            
        Returns:
            str: Column name or None
        """
        phone_keywords = ['number', 'phone', 'mobile', 'contact', 'tel', 'telephone', 'no.', 'no']
        
        for column in df.columns:
            col_lower = str(column).lower()
            
            for keyword in phone_keywords:
                if keyword in col_lower:
                    # Confirm with user
                    confirm = input(f"\n❓ Use column '{column}'? (yes/no): ").strip().lower()
                    if confirm in ['yes', 'y']:
                        return column
        
        return None
    
    @staticmethod
    def _extract_phone_numbers_from_column(ws, column_index: int) -> list:
        """
        Extract phone numbers from a column in Excel
        
        Args:
            ws: Worksheet object
            column_index: Column index
            
        Returns:
            list: List of phone numbers
        """
        numbers = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            value = row[column_index - 1]
            if value:
                number = str(value).strip()
                if number and number != 'None':
                    numbers.append(number)
        
        return numbers
