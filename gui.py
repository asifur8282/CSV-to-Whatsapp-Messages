"""
WhatsApp AutoMessenger - Tkinter GUI Interface
Provides a modern desktop interface for WhatsApp message automation.
"""

import sys
import os
import threading
import queue
import time
import random
import re
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
import pandas as pd

# Ensure the root project directory is in the path
project_root = str(Path(__file__).parent.absolute())
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Import CLI modules without modification
from src.message_sender import MessageSender
from src.file_handler import FileHandler
from src.user_input import UserInputHandler

# Color Palette (Slate & Emerald Dark Theme)
BG_COLOR = "#0f172a"          # Slate 900
CARD_BG = "#1e293b"           # Slate 800
TEXT_COLOR = "#f8fafc"        # Slate 50
TEXT_MUTED = "#94a3b8"        # Slate 400
BORDER_COLOR = "#334155"      # Slate 700
ACCENT_COLOR = "#10b981"      # Emerald 500
ACCENT_HOVER = "#059669"      # Emerald 600
RED_COLOR = "#ef4444"         # Red 500
RED_HOVER = "#dc2626"         # Red 600
CONSOLE_BG = "#020617"        # Slate 950


class QueueWriter:
    """Redirects writes to a queue for thread-safe UI updates."""
    def __init__(self, log_queue, original_stream):
        self.log_queue = log_queue
        self.original_stream = original_stream

    def write(self, text):
        self.log_queue.put(text)
        self.original_stream.write(text)
        self.original_stream.flush()

    def flush(self):
        self.original_stream.flush()


class RedirectStdout:
    """Context manager for redirecting stdout/stderr to a queue."""
    def __init__(self, log_queue):
        self.log_queue = log_queue
        self.old_stdout = sys.stdout
        self.old_stderr = sys.stderr

    def __enter__(self):
        sys.stdout = QueueWriter(self.log_queue, sys.stdout)
        sys.stderr = QueueWriter(self.log_queue, sys.stderr)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        sys.stdout = self.old_stdout
        sys.stderr = self.old_stderr


class AutoMessengerGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("WhatsApp AutoMessenger")
        self.geometry("1000x780")
        self.configure(bg=BG_COLOR)
        
        # Minimum size to ensure readability
        self.minsize(950, 700)
        
        # Threading & Control Variables
        self.automation_thread = None
        self.message_sender = None
        self.log_queue = queue.Queue()
        self.is_running = False
        
        # Form Data Variables
        self.source_var = tk.StringVar(value="file")
        self.file_path_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.column_var = tk.StringVar()
        self.limit_rows_var = tk.BooleanVar(value=False)
        self.limit_value_var = tk.StringVar(value="30")
        
        self.manual_country_var = tk.StringVar(value="91")
        self.manual_phone_var = tk.StringVar()
        self.manual_numbers = [] # list of collected phone numbers
        
        # Configure overall style
        self.setup_styles()
        
        # Build UI layout
        self.create_widgets()
        
        # Start queue polling loop
        self.poll_log_queue()
        
        # Bind window close event
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_styles(self):
        """Set up modern theme colors and widget styling using ttk."""
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # Main frames styling
        self.style.configure(".", background=BG_COLOR, foreground=TEXT_COLOR)
        self.style.configure("TFrame", background=BG_COLOR)
        self.style.configure("Card.TFrame", background=CARD_BG, borderwidth=1, relief="solid")
        
        # Labels
        self.style.configure("TLabel", background=BG_COLOR, foreground=TEXT_COLOR, font=("Segoe UI", 10))
        self.style.configure("Card.TLabel", background=CARD_BG, foreground=TEXT_COLOR, font=("Segoe UI", 10))
        self.style.configure("Title.TLabel", background=BG_COLOR, foreground=ACCENT_COLOR, font=("Segoe UI", 16, "bold"))
        self.style.configure("Header.TLabel", background=CARD_BG, foreground=ACCENT_COLOR, font=("Segoe UI", 11, "bold"))
        
        # Notebook (Tabs)
        self.style.configure("TNotebook", background=BG_COLOR, borderwidth=0)
        self.style.configure("TNotebook.Tab", background=CARD_BG, foreground=TEXT_MUTED, font=("Segoe UI", 10, "bold"), padding=[15, 5])
        self.style.map("TNotebook.Tab",
            background=[("selected", ACCENT_COLOR)],
            foreground=[("selected", BG_COLOR)]
        )
        
        # Combobox
        self.style.configure("TCombobox", fieldbackground=BG_COLOR, background=CARD_BG, foreground=TEXT_COLOR, arrowcolor=TEXT_COLOR)
        
        # Checkbutton
        self.style.configure("TCheckbutton", background=CARD_BG, foreground=TEXT_COLOR, font=("Segoe UI", 10))
        self.style.map("TCheckbutton",
            background=[("active", CARD_BG)],
            foreground=[("active", TEXT_COLOR)]
        )
        
        # Radiobutton
        self.style.configure("TRadiobutton", background=CARD_BG, foreground=TEXT_COLOR, font=("Segoe UI", 10))
        self.style.map("TRadiobutton",
            background=[("active", CARD_BG)],
            foreground=[("active", TEXT_COLOR)]
        )

    def create_widgets(self):
        # Top Header Banner
        header_frame = ttk.Frame(self)
        header_frame.pack(fill="x", padx=20, pady=15)
        
        title_lbl = ttk.Label(header_frame, text="🚀 WhatsApp AutoMessenger", style="Title.TLabel")
        title_lbl.pack(side="left")
        
        subtitle_lbl = ttk.Label(header_frame, text="Desktop Automation Tool", font=("Segoe UI", 10, "italic"), foreground=TEXT_MUTED)
        subtitle_lbl.pack(side="left", padx=10, pady=5)
        
        # Main Split Container
        main_pane = ttk.Frame(self)
        main_pane.pack(fill="both", expand=True, padx=20, pady=5)
        
        # Left Panel (Inputs)
        left_panel = ttk.Frame(main_pane)
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 10))
        
        # Right Panel (Controls & Logs)
        right_panel = ttk.Frame(main_pane)
        right_panel.pack(side="right", fill="both", expand=True, padx=(10, 0))
        
        # --- LEFT PANEL CONTENTS ---
        
        # 1. Source Selector Card
        source_card = ttk.Frame(left_panel, style="Card.TFrame", padding=15)
        source_card.pack(fill="x", pady=(0, 15))
        
        source_title = ttk.Label(source_card, text="1. Select Recipient Source", style="Header.TLabel")
        source_title.pack(anchor="w", pady=(0, 10))
        
        source_choices = ttk.Frame(source_card, style="Card.TFrame")
        source_choices.pack(fill="x")
        
        file_radio = ttk.Radiobutton(source_choices, text="Excel / CSV Spreadsheet", variable=self.source_var, value="file", command=self.toggle_source_mode)
        file_radio.pack(side="left", padx=(0, 20))
        
        manual_radio = ttk.Radiobutton(source_choices, text="Manual Numbers List", variable=self.source_var, value="manual", command=self.toggle_source_mode)
        manual_radio.pack(side="left")
        
        # 2. Dynamic Source Config Cards Container
        self.config_container = ttk.Frame(left_panel)
        self.config_container.pack(fill="both", expand=True)
        
        # 2a. Spreadsheet Mode Card
        self.file_card = ttk.Frame(self.config_container, style="Card.TFrame", padding=15)
        
        file_title = ttk.Label(self.file_card, text="2. Spreadsheet Configuration", style="Header.TLabel")
        file_title.pack(anchor="w", pady=(0, 10))
        
        # File selector row
        file_row = ttk.Frame(self.file_card, style="Card.TFrame")
        file_row.pack(fill="x", pady=5)
        
        self.file_path_entry = tk.Entry(file_row, textvariable=self.file_path_var, bg=BG_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR, relief="flat", highlightthickness=1, highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR, font=("Segoe UI", 10))
        self.file_path_entry.pack(side="left", fill="x", expand=True, padx=(0, 10), ipady=4)
        
        browse_btn = tk.Button(file_row, text="Browse File", bg=ACCENT_COLOR, fg=BG_COLOR, activebackground=ACCENT_HOVER, activeforeground=BG_COLOR, font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", padx=10, command=self.browse_file)
        browse_btn.pack(side="right")
        
        # Spreadsheet Parsing Options (comboboxes populated dynamically)
        self.excel_options_frame = ttk.Frame(self.file_card, style="Card.TFrame")
        self.excel_options_frame.pack(fill="x", pady=10)
        
        # Sheet combobox
        sheet_lbl = ttk.Label(self.excel_options_frame, text="Sheet Name:", style="Card.TLabel")
        sheet_lbl.grid(row=0, column=0, sticky="w", pady=5, padx=(0, 10))
        self.sheet_combo = ttk.Combobox(self.excel_options_frame, textvariable=self.sheet_var, state="readonly")
        self.sheet_combo.grid(row=0, column=1, sticky="ew", pady=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)
        
        # Column combobox
        col_lbl = ttk.Label(self.excel_options_frame, text="Phone Column:", style="Card.TLabel")
        col_lbl.grid(row=1, column=0, sticky="w", pady=5, padx=(0, 10))
        self.col_combo = ttk.Combobox(self.excel_options_frame, textvariable=self.column_var, state="readonly")
        self.col_combo.grid(row=1, column=1, sticky="ew", pady=5)
        
        self.excel_options_frame.columnconfigure(1, weight=1)
        
        # Limit rows limit row
        limit_row = ttk.Frame(self.file_card, style="Card.TFrame")
        limit_row.pack(fill="x", pady=5)
        
        limit_chk = ttk.Checkbutton(limit_row, text="Limit number of contacts to send to", variable=self.limit_rows_var, command=self.toggle_limit_entry)
        limit_chk.pack(side="left")
        
        self.limit_entry = tk.Entry(limit_row, textvariable=self.limit_value_var, width=6, bg=BG_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR, relief="flat", highlightthickness=1, highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR, state="disabled", font=("Segoe UI", 10), justify="center")
        self.limit_entry.pack(side="left", padx=10, ipady=2)
        
        # 2b. Manual Mode Card
        self.manual_card = ttk.Frame(self.config_container, style="Card.TFrame", padding=15)
        
        manual_title = ttk.Label(self.manual_card, text="2. Manual Numbers List", style="Header.TLabel")
        manual_title.pack(anchor="w", pady=(0, 10))
        
        # Number entry row (Country code + Phone number)
        entry_row = ttk.Frame(self.manual_card, style="Card.TFrame")
        entry_row.pack(fill="x", pady=5)
        
        cc_lbl = ttk.Label(entry_row, text="Country Code:", style="Card.TLabel")
        cc_lbl.pack(side="left", padx=(0, 5))
        
        cc_entry = tk.Entry(entry_row, textvariable=self.manual_country_var, width=5, bg=BG_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR, relief="flat", highlightthickness=1, highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR, font=("Segoe UI", 10), justify="center")
        cc_entry.pack(side="left", padx=(0, 15), ipady=4)
        
        num_lbl = ttk.Label(entry_row, text="Phone Number:", style="Card.TLabel")
        num_lbl.pack(side="left", padx=(0, 5))
        
        self.num_entry = tk.Entry(entry_row, textvariable=self.manual_phone_var, bg=BG_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR, relief="flat", highlightthickness=1, highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR, font=("Segoe UI", 10))
        self.num_entry.pack(side="left", fill="x", expand=True, padx=(0, 10), ipady=4)
        self.num_entry.bind("<Return>", lambda e: self.add_manual_number())
        
        add_btn = tk.Button(entry_row, text="Add", bg=ACCENT_COLOR, fg=BG_COLOR, activebackground=ACCENT_HOVER, activeforeground=BG_COLOR, font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", padx=15, command=self.add_manual_number)
        add_btn.pack(side="right")
        
        # List of added numbers
        list_frame = ttk.Frame(self.manual_card, style="Card.TFrame")
        list_frame.pack(fill="both", expand=True, pady=10)
        
        list_scroll = tk.Scrollbar(list_frame, orient="vertical")
        list_scroll.pack(side="right", fill="y")
        
        self.numbers_listbox = tk.Listbox(list_frame, bg=BG_COLOR, fg=TEXT_COLOR, selectbackground=ACCENT_COLOR, selectforeground=BG_COLOR, borderwidth=0, highlightthickness=1, highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR, font=("Consolas", 10), yscrollcommand=list_scroll.set)
        self.numbers_listbox.pack(side="left", fill="both", expand=True)
        list_scroll.config(command=self.numbers_listbox.yview)
        
        list_buttons = ttk.Frame(self.manual_card, style="Card.TFrame")
        list_buttons.pack(fill="x")
        
        self.list_count_lbl = ttk.Label(list_buttons, text="Total Recipients: 0", style="Card.TLabel", foreground=TEXT_MUTED)
        self.list_count_lbl.pack(side="left")
        
        remove_btn = tk.Button(list_buttons, text="Remove Selected", bg=RED_COLOR, fg=TEXT_COLOR, activebackground=RED_HOVER, activeforeground=TEXT_COLOR, font=("Segoe UI", 8, "bold"), relief="flat", cursor="hand2", padx=10, command=self.remove_manual_number)
        remove_btn.pack(side="right")
        
        # Pack the default card
        self.file_card.pack(fill="both", expand=True)
        
        # 3. Message Area Card
        msg_card = ttk.Frame(left_panel, style="Card.TFrame", padding=15)
        msg_card.pack(fill="both", expand=True, pady=(15, 0))
        
        msg_header = ttk.Frame(msg_card, style="Card.TFrame")
        msg_header.pack(fill="x", pady=(0, 5))
        
        msg_title = ttk.Label(msg_header, text="3. Message Compose", style="Header.TLabel")
        msg_title.pack(side="left")
        
        self.char_count_lbl = ttk.Label(msg_header, text="Characters: 0/4096", style="Card.TLabel", foreground=TEXT_MUTED)
        self.char_count_lbl.pack(side="right")
        
        self.msg_text = scrolledtext.ScrolledText(msg_card, bg=BG_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR, borderwidth=0, highlightthickness=1, highlightbackground=BORDER_COLOR, highlightcolor=ACCENT_COLOR, font=("Segoe UI", 10), wrap="word", height=8)
        self.msg_text.pack(fill="both", expand=True)
        self.msg_text.bind("<KeyRelease>", self.update_char_count)
        
        # --- RIGHT PANEL CONTENTS ---
        
        # 4. Actions & Progress Card
        actions_card = ttk.Frame(right_panel, style="Card.TFrame", padding=15)
        actions_card.pack(fill="x", pady=(0, 15))
        
        actions_title = ttk.Label(actions_card, text="Execution Controls", style="Header.TLabel")
        actions_title.pack(anchor="w", pady=(0, 10))
        
        btn_row = ttk.Frame(actions_card, style="Card.TFrame")
        btn_row.pack(fill="x", pady=5)
        
        self.start_btn = tk.Button(btn_row, text="Start Automating", bg=ACCENT_COLOR, fg=BG_COLOR, activebackground=ACCENT_HOVER, activeforeground=BG_COLOR, font=("Segoe UI", 11, "bold"), relief="flat", cursor="hand2", padx=20, pady=8, command=self.start_automation)
        self.start_btn.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        self.stop_btn = tk.Button(btn_row, text="Stop / Close Browser", bg=RED_COLOR, fg=TEXT_COLOR, activebackground=RED_HOVER, activeforeground=TEXT_COLOR, font=("Segoe UI", 11, "bold"), relief="flat", state="disabled", cursor="hand2", padx=20, pady=8, command=self.stop_automation)
        self.stop_btn.pack(side="right", fill="x", expand=True, padx=(10, 0))
        
        progress_row = ttk.Frame(actions_card, style="Card.TFrame")
        progress_row.pack(fill="x", pady=(15, 0))
        
        self.progress_lbl = ttk.Label(progress_row, text="Status: Ready to start", style="Card.TLabel", foreground=TEXT_MUTED)
        self.progress_lbl.pack(anchor="w", pady=(0, 5))
        
        self.progress_bar = ttk.Progressbar(progress_row, mode="determinate")
        self.progress_bar.pack(fill="x")
        
        # 5. Live Console Card
        console_card = ttk.Frame(right_panel, style="Card.TFrame", padding=15)
        console_card.pack(fill="both", expand=True)
        
        console_title = ttk.Label(console_card, text="Execution Logs", style="Header.TLabel")
        console_title.pack(anchor="w", pady=(0, 10))
        
        self.log_text = scrolledtext.ScrolledText(console_card, bg=CONSOLE_BG, fg="#10b981", insertbackground="#10b981", borderwidth=0, font=("Consolas", 9), wrap="word")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.config(state="disabled")

    # --- UI EVENT HANDLERS ---
    
    def toggle_source_mode(self):
        """Show the selected config frame and hide the other."""
        mode = self.source_var.get()
        if mode == "file":
            self.manual_card.pack_forget()
            self.file_card.pack(fill="both", expand=True)
        else:
            self.file_card.pack_forget()
            self.manual_card.pack(fill="both", expand=True)

    def toggle_limit_entry(self):
        """Enable/Disable limit entry based on checkbox."""
        if self.limit_rows_var.get():
            self.limit_entry.config(state="normal")
        else:
            self.limit_entry.config(state="disabled")

    def browse_file(self):
        """Open file chooser dialog and update file path, sheet names, and columns."""
        file_path = filedialog.askopenfilename(
            title="Select Spreadsheet File",
            filetypes=[("Spreadsheet Files", "*.xlsx *.xls *.csv"), ("Excel Files", "*.xlsx *.xls"), ("CSV Files", "*.csv")]
        )
        if not file_path:
            return
            
        self.file_path_var.set(file_path)
        
        # Reset dropdown lists
        self.sheet_combo.config(values=[])
        self.sheet_var.set("")
        self.col_combo.config(values=[])
        self.column_var.set("")
        
        # Parse sheet and column names based on file type
        try:
            if file_path.lower().endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                
                if not sheets:
                    messagebox.showerror("Error", "Excel file contains no worksheets.")
                    return
                
                self.sheet_combo.config(values=sheets)
                self.sheet_var.set(sheets[0])
                self.on_sheet_selected(None) # Automatically trigger column detection for first sheet
            else:
                # CSV File
                df = pd.read_csv(file_path, nrows=1)
                columns = list(df.columns)
                
                self.sheet_combo.config(values=["N/A (CSV file)"])
                self.sheet_var.set("N/A (CSV file)")
                self.col_combo.config(values=columns)
                
                # Try to auto-select column matching 'number' keywords
                phone_keywords = ['number', 'phone', 'mobile', 'contact', 'tel', 'telephone', 'no.', 'no']
                for col in columns:
                    if any(keyword in str(col).lower() for keyword in phone_keywords):
                        self.column_var.set(col)
                        break
                if not self.column_var.get() and columns:
                    self.column_var.set(columns[0])
        except Exception as e:
            messagebox.showerror("File Error", f"Failed to parse selected file:\n{str(e)}")

    def on_sheet_selected(self, event):
        """Read column names from Excel worksheet when selected."""
        file_path = self.file_path_var.get()
        sheet_name = self.sheet_var.get()
        if not file_path or not sheet_name or sheet_name == "N/A (CSV file)":
            return
            
        try:
            # We open the worksheet to read the first row
            wb = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                first_row = next(ws.iter_rows(max_row=1, values_only=True))
                columns = [str(cell).strip() for cell in first_row if cell is not None]
                
                self.col_combo.config(values=columns)
                
                # Auto-detect column
                phone_keywords = ['number', 'phone', 'mobile', 'contact', 'tel', 'telephone', 'no.', 'no']
                for col in columns:
                    if any(keyword in str(col).lower() for keyword in phone_keywords):
                        self.column_var.set(col)
                        break
                if not self.column_var.get() and columns:
                    self.column_var.set(columns[0])
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read sheet columns:\n{str(e)}")

    def add_manual_number(self):
        """Validate and add number from text entry fields to listbox."""
        country_code = self.manual_country_var.get().strip().replace("+", "")
        phone_number = self.manual_phone_var.get().strip().replace(" ", "").replace("-", "")
        
        if not country_code:
            messagebox.showwarning("Validation Warning", "Please enter a country code (e.g. 91).")
            return
            
        if not phone_number:
            messagebox.showwarning("Validation Warning", "Please enter a phone number.")
            return
            
        # Combine country code + number
        full_number = f"{country_code}{phone_number}"
        
        # Check against base validation method in UserInputHandler
        if UserInputHandler._validate_phone_number(full_number):
            if full_number in self.manual_numbers:
                messagebox.showinfo("Duplicate", "This number is already in the recipient list.")
                return
            
            self.manual_numbers.append(full_number)
            self.numbers_listbox.insert(tk.END, full_number)
            self.manual_phone_var.set("") # Clear phone input
            self.update_list_count()
        else:
            messagebox.showwarning("Invalid Number", f"The number '{full_number}' is invalid.\nEnsure it contains only digits (7 to 15 digits total).")

    def remove_manual_number(self):
        """Remove selected number from listbox."""
        selected_indices = self.numbers_listbox.curselection()
        if not selected_indices:
            messagebox.showinfo("Selection", "Please select a number from the list to remove.")
            return
            
        # Remove in reverse order to preserve correct indexes
        for index in sorted(selected_indices, reverse=True):
            self.manual_numbers.pop(index)
            self.numbers_listbox.delete(index)
            
        self.update_list_count()

    def update_list_count(self):
        """Update count label for manual numbers list."""
        self.list_count_lbl.config(text=f"Total Recipients: {len(self.manual_numbers)}")

    def update_char_count(self, event=None):
        """Update live character count label for message compose."""
        message = self.msg_text.get("1.0", tk.END).strip()
        length = len(message)
        self.char_count_lbl.config(text=f"Characters: {length}/4096")
        if length > 4096:
            self.char_count_lbl.config(foreground=RED_COLOR)
        else:
            self.char_count_lbl.config(foreground=TEXT_MUTED)

    # --- LOGGER & QUEUE POLLING ---

    def log(self, message):
        """Log a message to the Tkinter scroll widget."""
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, message)
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")

    def poll_log_queue(self):
        """Poll the queue periodically for any background log writes."""
        while not self.log_queue.empty():
            try:
                log_data = self.log_queue.get_nowait()
                self.log(log_data)
                self.update_progress_status(log_data)
            except queue.Empty:
                break
                
        # Repeat every 100ms
        self.after(100, self.poll_log_queue)

    def update_progress_status(self, log_data):
        """Parse status updates from the standard print output strings and update GUI metrics."""
        # Detect recipient loop
        # Format: "Sending to recipient 2/10: 919876543210"
        match_progress = re.search(r"Sending to recipient (\d+)/(\d+):", log_data)
        if match_progress:
            current = int(match_progress.group(1))
            total = int(match_progress.group(2))
            self.progress_lbl.config(text=f"Status: Sending messages ({current}/{total})...")
            self.progress_bar.config(mode="determinate", maximum=total, value=current - 0.5)
            return

        # Detect human wait timer
        # Format: "Waiting for 15 seconds before the next message..."
        match_wait = re.search(r"Waiting for (\d+) seconds", log_data)
        if match_wait:
            sec = match_wait.group(1)
            self.progress_lbl.config(text=f"Status: Waiting {sec}s to mimic human activity...")
            return

        # Detect QR Code scan wait
        # Format: "Waiting for you to login"
        if "Waiting for you to login" in log_data:
            self.progress_lbl.config(text="Status: Please scan QR code in the browser!")
            self.progress_bar.config(mode="indeterminate")
            self.progress_bar.start(10)
            return

        # Detect successful login
        if "Login successful" in log_data:
            self.progress_bar.stop()
            self.progress_bar.config(mode="determinate", value=0)
            self.progress_lbl.config(text="Status: Login successful! Sending...")
            return

        # Detect end of processing
        if "MESSAGING SUMMARY" in log_data or "All messages sent successfully" in log_data or "No messages were sent successfully" in log_data:
            self.progress_bar.stop()
            self.progress_lbl.config(text="Status: Process completed!")
            return

    # --- AUTOMATION WORKER THREAD ---

    def start_automation(self):
        """Validate inputs and launch the automation thread."""
        if self.is_running:
            return
            
        # Get and validate message
        message = self.msg_text.get("1.0", tk.END).strip()
        if not message:
            messagebox.showwarning("Missing Input", "Message cannot be empty!")
            return
            
        if len(message) > 4096:
            messagebox.showerror("Input Error", "Message exceeds maximum allowed length (4096 characters).")
            return
            
        # Gather recipients
        mode = self.source_var.get()
        recipients = []
        
        if mode == "file":
            file_path = self.file_path_var.get()
            sheet_name = self.sheet_var.get()
            column_name = self.column_var.get()
            
            if not file_path:
                messagebox.showwarning("Missing Input", "Please select an Excel or CSV file.")
                return
            if not column_name:
                messagebox.showwarning("Missing Input", "Please select the phone number column.")
                return
                
            # Extract phone numbers from file
            try:
                if file_path.lower().endswith(('.xlsx', '.xls')):
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    ws = wb[sheet_name]
                    
                    # Find column index
                    col_index = None
                    first_row = next(ws.iter_rows(max_row=1, values_only=True))
                    for idx, val in enumerate(first_row, 1):
                        if str(val).strip() == column_name:
                            col_index = idx
                            break
                            
                    if col_index is None:
                        messagebox.showerror("Error", f"Could not find column '{column_name}' in sheet.")
                        return
                        
                    # Extract values
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        val = row[col_index - 1]
                        if val:
                            num = str(val).strip()
                            if num and num != 'None':
                                recipients.append(num)
                    wb.close()
                else:
                    # CSV file
                    df = pd.read_csv(file_path)
                    recipients = [str(num).strip() for num in df[column_name].astype(str) if str(num).strip()]
            except Exception as e:
                messagebox.showerror("File Parsing Error", f"Could not extract numbers from spreadsheet:\n{str(e)}")
                return
                
            if not recipients:
                messagebox.showwarning("No Data", "No recipients found in the selected sheet/column.")
                return
                
            # Apply row limits
            if self.limit_rows_var.get():
                try:
                    limit = int(self.limit_value_var.get())
                    if limit < 1:
                        raise ValueError()
                    recipients = recipients[:limit]
                except ValueError:
                    messagebox.showwarning("Validation Warning", "Invalid row limit. Please enter a valid number >= 1.")
                    return
        else:
            # Manual recipients list
            recipients = self.manual_numbers.copy()
            if not recipients:
                messagebox.showwarning("Missing Input", "Please add at least one recipient to the list.")
                return

        # Double check recipients count safety
        if len(recipients) > 30:
            proceed = messagebox.askyesno(
                "Large Batch Warning",
                f"You are attempting to send messages to {len(recipients)} contacts.\n"
                "Sending messages to more than 30 contacts increases the risk of a WhatsApp ban.\n\n"
                "Do you want to proceed?"
            )
            if not proceed:
                return

        # Check if browser is already open and alive from a previous run
        browser_alive = False
        if self.message_sender and self.message_sender.bot:
            driver = self.message_sender.bot.driver
            if driver:
                try:
                    # Accessing the title verifies if the browser process is still running and responsive
                    _ = driver.title
                    browser_alive = True
                except Exception:
                    pass

        if not browser_alive:
            # Clean up old inactive session if any exists
            if self.message_sender:
                try:
                    self.message_sender.close()
                except:
                    pass
            # Create a fresh MessageSender instance
            self.message_sender = MessageSender()
        else:
            self.log("\n🔄 Reusing existing active browser session...\n")

        # Prepare GUI state for running
        self.is_running = True
        self.start_btn.config(state="disabled", text="Running...")
        self.stop_btn.config(state="normal")
        self.progress_bar.config(value=0)
        self.progress_lbl.config(text="Status: Initializing browser..." if not browser_alive else "Status: Reusing browser...")
        
        # Clear log console
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.config(state="disabled")
        
        # Start Thread
        self.automation_thread = threading.Thread(
            target=self.run_automation_loop,
            args=(recipients, message),
            daemon=True
        )
        self.automation_thread.start()

    def run_automation_loop(self, recipients, message):
        """Worker thread entry point."""
        # Use RedirectStdout context manager to capture print output from standard library
        with RedirectStdout(self.log_queue):
            try:
                # Runs the exact CLI sender engine logic
                self.message_sender.send_messages(recipients, message)
            except Exception as e:
                print(f"\n❌ Thread Exception: {str(e)}")
            finally:
                # Post completion event to UI thread
                self.after(0, self.on_automation_complete)

    def on_automation_complete(self):
        """Update GUI back to standby state when thread finishes."""
        self.is_running = False
        self.start_btn.config(state="normal", text="Start Automating")
        self.stop_btn.config(state="normal")  # Keep stop button active to close the browser
        self.progress_bar.stop()
        self.progress_bar.config(mode="determinate", value=0)
        self.progress_lbl.config(text="Status: Completed / Standby")
        self.automation_thread = None

    def stop_automation(self):
        """Close the web browser or abort active automation."""
        if not self.message_sender:
            return
            
        if self.is_running:
            confirm = messagebox.askyesno(
                "Confirm Abort",
                "Are you sure you want to stop the automation process and close the WhatsApp Web window?"
            )
            if not confirm:
                return
            self.log("\n⚠️  Stopping automation... Closing Chrome Driver...\n")
            self.progress_lbl.config(text="Status: Aborting...")
        else:
            confirm = messagebox.askyesno(
                "Close Browser",
                "Do you want to close the active WhatsApp Web browser window?"
            )
            if not confirm:
                return
            self.log("\n⚠️  Closing Chrome Driver...\n")
            
        try:
            self.message_sender.close()
            self.stop_btn.config(state="disabled")
        except Exception as e:
            self.log(f"Error closing browser: {str(e)}\n")

    def on_closing(self):
        """Ensure everything is cleaned up properly when closing window."""
        if self.is_running:
            if not messagebox.askyesno("Quit Application", "Automation is currently running. Do you want to force quit?"):
                return
            
            # Attempt clean terminate
            if self.message_sender:
                try:
                    self.message_sender.close()
                except:
                    pass
        
        self.destroy()


if __name__ == "__main__":
    app = AutoMessengerGUI()
    app.mainloop()
